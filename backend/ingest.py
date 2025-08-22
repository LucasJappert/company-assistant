import os
import uuid
from pathlib import Path
from typing import List, Dict, Union, Mapping, cast

import numpy as np  # <-- NUEVO
from chromadb.api.types import EmbeddingFunction, Documents, Embeddings


class DummyEmbeddingFunction(EmbeddingFunction):
    def __call__(self, input: Documents) -> Embeddings:
        # Devolvemos un vector “dummy” por cada documento de entrada.
        # No se usa realmente porque pasamos embeddings manuales en .add(...)
        return [[0.0] for _ in input]


# ChromaDB como vector store local (sin servidor)
import chromadb

from dotenv import load_dotenv

# Embeddings: BGE-M3 (gratuito y multilingüe)
from FlagEmbedding import BGEM3FlagModel

# Unstructured para extraer texto de múltiples formatos
from chromadb.config import Settings

load_dotenv()

CHROMA_DIR = os.getenv("CHROMA_DIR", "../chroma_storage")
CHROMA_PATH = str((Path(__file__).resolve().parent / CHROMA_DIR))
COLLECTION = "company_docs"
DATA_DIR = Path(__file__).resolve().parent.parent / "data" / "raw"

# OCR opcional
TESSERACT_CMD = os.getenv("TESSERACT_CMD")
if TESSERACT_CMD:
    import pytesseract

    pytesseract.pytesseract.tesseract_cmd = TESSERACT_CMD

# BGE-M3: 1024 dims
embed_model = BGEM3FlagModel("BAAI/bge-m3", use_fp16=True)


# Inicializa Chroma con persistencia en disco
client = chromadb.PersistentClient(path=CHROMA_PATH, settings=Settings())
collection = client.get_or_create_collection(
    name=COLLECTION,
    embedding_function=DummyEmbeddingFunction(),
)


def chunk_text(txt: str, chunk_size: int = 1500, overlap: int = 200) -> List[str]:
    txt = " ".join(txt.split())  # normaliza espacios
    chunks: List[str] = []
    start = 0
    while start < len(txt):
        end = min(len(txt), start + chunk_size)
        chunks.append(txt[start:end])
        if end == len(txt):
            break
        start = max(0, end - overlap)
    return chunks


def extract_text_from_file(path: Path) -> str:
    ext = path.suffix.lower()
    try:
        if ext == ".pdf":
            from pdfminer.high_level import extract_text as pdf_extract_text

            return pdf_extract_text(str(path)) or ""

        elif ext == ".docx":
            from docx import Document

            doc = Document(str(path))
            return "\n".join(p.text for p in doc.paragraphs if p.text) or ""

        elif ext == ".xlsx":
            from openpyxl import load_workbook

            wb = load_workbook(filename=str(path), read_only=True, data_only=True)
            out_lines = []
            for ws in wb.worksheets:
                out_lines.append(f"# Hoja: {ws.title}")
                for row in ws.iter_rows(values_only=True):
                    vals = [str(v) for v in row if v is not None]
                    if vals:
                        out_lines.append(" | ".join(vals))
            wb.close()
            return "\n".join(out_lines)

        elif ext == ".csv":
            import csv

            out_lines = []
            with open(path, "r", encoding="utf-8", errors="ignore", newline="") as f:
                reader = csv.reader(f)
                for row in reader:
                    out_lines.append(", ".join(row))
            return "\n".join(out_lines)

        elif ext in (".txt", ".md"):
            return path.read_text(encoding="utf-8", errors="ignore")

        else:
            print(f"[WARN] Unsupported file type: {ext} ({path.name})")
            return ""

    except Exception as ex:
        print(f"[WARN] Extract failed for {path.name}: {ex}")
        return ""


def gather_documents() -> List[Dict[str, str]]:
    docs: List[Dict[str, str]] = []
    for p in DATA_DIR.rglob("*"):
        if p.is_file() and p.suffix.lower() in {
            ".pdf",
            ".docx",
            ".xlsx",
            ".txt",
            ".md",
            ".csv",
            ".pptx",
        }:
            try:
                text = extract_text_from_file(p)
                if text.strip():
                    docs.append({"path": str(p), "text": text})
            except Exception as ex:
                print(f"[WARN] Skipping {p.name}: {ex}")
    return docs


def embed_texts(texts: List[str]) -> List[List[float]]:
    out = embed_model.encode(texts, batch_size=32)
    dense = out["dense_vecs"]  # type: ignore[index]
    # Aseguramos List[List[float]] (no ndarray)
    arr = np.asarray(dense, dtype=np.float32)
    return cast(List[List[float]], arr.tolist())


# Tipado de metadatos que espera Chroma
MetaValue = Union[str, int, float, bool]
Metadata = Mapping[str, MetaValue]  # para las anotaciones de entrada a .add


def main() -> None:
    raw_docs = gather_documents()
    print(f"Found {len(raw_docs)} document(s).")

    ids: List[str] = []
    documents: List[str] = []
    metadatas: List[Dict[str, MetaValue]] = (
        []
    )  # list de dicts simples str->(tipos primitivos)

    for d in raw_docs:
        chunks = chunk_text(d["text"])
        for idx, ch in enumerate(chunks):
            ids.append(str(uuid.uuid4()))
            documents.append(ch)
            metadatas.append({"source": d["path"], "chunk_idx": idx})

    print(f"Prepared {len(documents)} chunk(s). Embedding & upserting...")
    embeddings_ll = embed_texts(documents)  # List[List[float]]
    embeddings_np = np.asarray(
        embeddings_ll, dtype=np.float32
    )  # <-- convierto a ndarray

    collection.add(
        ids=ids,
        documents=documents,
        metadatas=cast(List[Metadata], metadatas),
        embeddings=embeddings_np,  # <-- ndarray: Pyright feliz y Chroma también
    )
    print("Done.")


if __name__ == "__main__":
    main()
